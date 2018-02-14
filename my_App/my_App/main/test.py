from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from cStringIO import StringIO
from abbreviations import us_state_abbrev

from openpyxl import load_workbook 

def performScrape(path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = file(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
        interpreter.process_page(page)

    outputMain = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()

    output = outputMain.split()

    itemsInOutPut = len(output)


    for i in range(itemsInOutPut):
        currentWord = output[i]
        if(currentWord == "Area:"):
            toSlice = output[i + 1]
            Areas.append(toSlice[-1])

        if(currentWord == "Rate"):
            for j in range(1,16):
                Rates.append(output[i + j])          

        if(currentWord == "Dates:"):
            startDates.append(output[i +1])
            endDates.append(output[i + 3])
            States.append(us_state_abbrev.get(str.lower(output[i + 4])))

        if(currentWord == "Name:"):
            NameToAdd = ""
            j = i +1
            while(output[j] != "Age"):
                NameToAdd += output[j] + " "
                j+=1
            planNames.append(NameToAdd[:-1])

   




 
    



Rates = []
startDates = []
endDates = []
Areas = []
planNames = []
States = []


for i in range(1,10):
    if( i != 4):
        performScrape("para0" + str(i) + ".pdf")



# for i in range(1,10):
#     if(i != 4):
#         performScrape('para0' + str(i) + ".pdf")



excelFile = load_workbook("BeneFix Small Group Plans upload template.xlsx")

mainSheet = excelFile.active


totalItems = len(startDates)

rateCount = 0

for i in range(totalItems):
    mainSheet.cell(i + 2,1).value = startDates[i]
    mainSheet.cell(i + 2,2).value = endDates[i]
    mainSheet.cell(i + 2,3).value = planNames[i]
    mainSheet.cell(i + 2,4).value = States[i]
    mainSheet.cell(i + 2,5).value = Areas[i]
    mainSheet.cell(i + 2, 6).value = Rates[rateCount]
    for j in range(0,45):
        mainSheet.cell(i + 2, j + 7).value = Rates[j + rateCount]
    mainSheet.cell(i + 2,52).value = Rates[44 + rateCount]
    rateCount+=45





excelFile.save("BeneFix Small Group Plans upload template.xlsx")




