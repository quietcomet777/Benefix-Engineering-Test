# main class that handles the pdf scraping and excel population 

from pdfHandler import pdfHandler
from abbreviations import us_state_abbrev
from openpyxl import load_workbook 


class mainScraper(object):
	"""docstring for ClassName"""
	def __init__(self):
		super(mainScraper, self).__init__()
		self.Rates = []
		self.startDates = []
		self.endDates = []
		self.Areas = []
		self.planNames = []
		self.States = []

	def dataExtraction(self,path):
		pdfData = pdfHandler(path)

		total = len(pdfData)

		for i in range(total):
			currentWord = pdfData[i]
			if(currentWord == "Area:"):
				self.Areas.append(pdfData[i +1][-1])

			if(currentWord == "Rate"):
				for j in range(1,16):
					self.Rates.append(pdfData[i + j])

			if(currentWord == "Dates:"):
				self.startDates.append(pdfData[i + 1])
				self.endDates.append(pdfData[i + 3])
				self.States.append(us_state_abbrev.get(str.lower(pdfData[i + 4])))

			if(currentWord == "Name:"):
				NameToAdd = ""
				j = i + 1
				while(pdfData[j] != "Age"):
					NameToAdd += pdfData[j] + " "
					j += 1
				self.planNames.append(NameToAdd[:-1])
     

	def writeToExcel(self,path):
		excelFile = load_workbook(path)
		mainSheet = excelFile.active
		totalItems = len(self.Areas)
		rateCount = 0

		for i in range(totalItems):
			currentRow = i + 2
			mainSheet.cell(currentRow, 1).value = self.startDates[i]
			mainSheet.cell(currentRow, 2).value = self.endDates[i]
			mainSheet.cell(currentRow,3).value = self.planNames[i]
			mainSheet.cell(currentRow,4).value = self.States[i]
			mainSheet.cell(currentRow, 5).value = self.Areas[i]
			mainSheet.cell(currentRow, 6).value = self.Rates[rateCount]
			for j in range(0,45):
				mainSheet.cell(currentRow, j + 7).value = self.Rates[j + rateCount]
			mainSheet.cell(currentRow, 52).value = self.Rates[44 + rateCount]
			rateCount+=45

		excelFile.save(path)


